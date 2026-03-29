import os
import requests
import urllib.parse
import openpyxl
from bs4 import BeautifulSoup
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════
# 品牌清單（每次新專案只需要改這裡）
# ══════════════════════════════════════════════════════════
BRANDS = [
    {
        "name": "東區德",
        "website": "https://dio3212.kaik.io",
        "fb_page_id": "155381434648518",
    },
    {
        "name": "UC Training",
        "website": "https://uctraining.cc",
        "fb_page_id": "110845790371563",
    },
    {
        "name": "斑馬兩性",
        "website": "https://zebrapua.com",
        "fb_page_id": "",
    },
    {
        "name": "想想交友",
        "website": "https://flowgalaxy.tw",
        "fb_page_id": "102940652715715",
    },
    {
        "name": "戀愛生活圈",
        "website": "https://blog.uctraining.cc",
        "fb_page_id": "",
    },
]
# ══════════════════════════════════════════════════════════

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36"
}
LOGO_DIR = "logos"
os.makedirs(LOGO_DIR, exist_ok=True)


# ── 步驟 1：抓 OG Image ───────────────────────────────────
def fetch_og_image(brand):
    name = brand["name"]
    url = brand["website"]
    try:
        res = requests.get(url, headers=HEADERS, timeout=10)
        soup = BeautifulSoup(res.text, "html.parser")

        og = soup.find("meta", property="og:image")
        if og and og.get("content"):
            img_url = og["content"]
            return download_image(name, img_url)

        icon = soup.find("link", rel=lambda r: r and "icon" in r)
        if icon and icon.get("href"):
            icon_url = urllib.parse.urljoin(url, icon["href"])
            print(f"⚠️  {name}：無 OG image，改用 favicon")
            return download_image(name, icon_url)

    except Exception as e:
        print(f"❌ {name}：圖片抓取失敗 → {e}")
    return None


def download_image(name, img_url):
    try:
        ext = img_url.split(".")[-1].split("?")[0]
        if ext.lower() not in ["jpg", "jpeg", "png", "gif", "webp", "ico", "svg"]:
            ext = "png"
        path = f"{LOGO_DIR}/{name}.{ext}"
        res = requests.get(img_url, headers=HEADERS, timeout=10)
        with open(path, "wb") as f:
            f.write(res.content)
        print(f"✅ {name}：圖片已儲存 → {path}")
        return path
    except Exception as e:
        print(f"❌ {name}：圖片下載失敗 → {e}")
        return None


# ── 步驟 2：產生廣告庫連結 ────────────────────────────────
def make_ad_library_url(brand):
    base = "https://www.facebook.com/ads/library/"
    page_id = brand.get("fb_page_id", "")
    if page_id:
        params = {
            "active_status": "all",
            "ad_type": "all",
            "country": "TW",
            "view_all_page_id": page_id,
            "search_type": "page",
            "media_type": "all",
        }
    else:
        params = {
            "active_status": "all",
            "ad_type": "all",
            "country": "TW",
            "q": brand["name"],
            "search_type": "keyword_unordered",
            "media_type": "all",
        }
    return base + "?" + urllib.parse.urlencode(params, quote_via=urllib.parse.quote)


# ── 步驟 3：產生 Excel ────────────────────────────────────
def build_excel(brands, logo_paths):
    wb = openpyxl.Workbook()
    thin = Side(style='thin', color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    def fill(hex_color):
        return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

    # ── Sheet 1：玩家總覽 ──────────────────────────────────
    ws1 = wb.active
    ws1.title = "玩家總覽"

    headers = ["#", "品牌名稱", "官網", "FB 廣告庫", "Logo 路徑"]
    widths  = [5, 18, 30, 35, 30]

    ws1.row_dimensions[1].height = 25
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws1.cell(row=1, column=i, value=h)
        cell.font = Font(name='Arial', bold=True, color="FFFFFF", size=10)
        cell.fill = fill("1A1A2E")
        cell.alignment = center
        cell.border = border
        ws1.column_dimensions[get_column_letter(i)].width = w

    for idx, brand in enumerate(brands, start=1):
        row = idx + 1
        row_fill = fill("F5F5F5") if idx % 2 == 0 else fill("FFFFFF")
        ad_url = make_ad_library_url(brand)
        logo_path = logo_paths.get(brand["name"], "未找到")

        for col, val in enumerate(
            [idx, brand["name"], brand["website"], "點我開啟廣告庫", logo_path],
            start=1
        ):
            cell = ws1.cell(row=row, column=col, value=val)
            cell.font = Font(name='Arial', size=9)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = left_align if col == 3 else center

        link_cell = ws1.cell(row=row, column=4)
        link_cell.hyperlink = ad_url
        link_cell.font = Font(name='Arial', size=9, color="1877F2", underline="single")

    ws1.freeze_panes = "A2"

    # ── Sheet 2：競品數據追蹤 ──────────────────────────────
    ws2 = wb.create_sheet("競品數據追蹤")

    columns = [
        ("品牌名稱", 16, "basic"),
        ("官網", 22, "basic"),
        ("FB 粉絲數", 12, "fb"),
        ("FB 近期均讚\n(最近10篇)", 14, "fb"),
        ("FB 近期均留言\n(最近10篇)", 14, "fb"),
        ("FB 近期互動率", 13, "fb"),
        ("FB 熱門均讚\n(最熱門10篇)", 14, "fb"),
        ("FB 熱門均留言\n(最熱門10篇)", 14, "fb"),
        ("FB 熱門互動率", 13, "fb"),
        ("IG 粉絲數", 12, "ig"),
        ("IG 近期均讚\n(最近10篇)", 14, "ig"),
        ("IG 近期均留言\n(最近10篇)", 14, "ig"),
        ("IG 近期互動率", 13, "ig"),
        ("IG 熱門均讚\n(最熱門10篇)", 14, "ig"),
        ("IG 熱門均留言\n(最熱門10篇)", 14, "ig"),
        ("IG 熱門互動率", 13, "ig"),
        ("YT 訂閱數", 12, "yt"),
        ("YT 近期均觀看\n(最近10篇)", 14, "yt"),
        ("YT 近期均讚\n(最近10篇)", 14, "yt"),
        ("YT 近期互動率\n(讚/觀看)", 14, "yt"),
        ("YT 熱門均觀看\n(最熱門10篇)", 14, "yt"),
        ("YT 熱門均讚\n(最熱門10篇)", 14, "yt"),
        ("YT 熱門互動率\n(讚/觀看)", 14, "yt"),
        ("有無投放廣告", 12, "biz"),
        ("廣告預算規模", 14, "biz"),
        ("財報／銷售額", 14, "biz"),
        ("備註", 20, "biz"),
    ]

    group_colors = {
        "basic": "1A1A2E",
        "fb":    "1877F2",
        "ig":    "C13584",
        "yt":    "FF0000",
        "biz":   "2D6A4F",
    }
    group_labels = {
        "basic": "基本資料", "fb": "Facebook",
        "ig": "Instagram", "yt": "YouTube", "biz": "商業規模",
    }

    group_ranges = {g: [] for g in group_colors}
    for idx, (_, _, group) in enumerate(columns, start=1):
        group_ranges[group].append(idx)

    ws2.row_dimensions[1].height = 22
    for group, cols in group_ranges.items():
        s, e = min(cols), max(cols)
        cell = ws2.cell(row=1, column=s, value=group_labels[group])
        cell.font = Font(name='Arial', bold=True, color="FFFFFF", size=11)
        cell.fill = fill(group_colors[group])
        cell.alignment = center
        cell.border = border
        if e > s:
            ws2.merge_cells(start_row=1, start_column=s, end_row=1, end_column=e)
            for c in range(s+1, e+1):
                ws2.cell(row=1, column=c).border = border

    ws2.row_dimensions[2].height = 40
    for idx, (header, width, group) in enumerate(columns, start=1):
        cell = ws2.cell(row=2, column=idx, value=header)
        cell.font = Font(name='Arial', bold=True, color="FFFFFF", size=9)
        cell.fill = fill(group_colors[group])
        cell.alignment = center
        cell.border = border
        ws2.column_dimensions[get_column_letter(idx)].width = width

    for row in range(3, 3 + len(brands)):
        brand = brands[row - 3]
        row_fill = fill("F5F5F5") if row % 2 == 0 else fill("FFFFFF")

        for col_idx in range(1, len(columns)+1):
            cell = ws2.cell(row=row, column=col_idx)
            cell.fill = row_fill
            cell.font = Font(name='Arial', size=9)
            cell.border = border

            if col_idx == 1:
                cell.value = brand["name"]
                cell.alignment = left_align
            elif col_idx == 2:
                cell.value = brand["website"]
                cell.alignment = left_align
            elif col_idx == 6:
                cell.value = f"=IFERROR((D{row}+E{row})/C{row},\"–\")"
                cell.alignment = center; cell.number_format = '0.00%'
            elif col_idx == 9:
                cell.value = f"=IFERROR((G{row}+H{row})/C{row},\"–\")"
                cell.alignment = center; cell.number_format = '0.00%'
            elif col_idx == 13:
                cell.value = f"=IFERROR((K{row}+L{row})/J{row},\"–\")"
                cell.alignment = center; cell.number_format = '0.00%'
            elif col_idx == 16:
                cell.value = f"=IFERROR((N{row}+O{row})/J{row},\"–\")"
                cell.alignment = center; cell.number_format = '0.00%'
            elif col_idx == 20:
                cell.value = f"=IFERROR(S{row}/R{row},\"–\")"
                cell.alignment = center; cell.number_format = '0.00%'
            elif col_idx == 23:
                cell.value = f"=IFERROR(V{row}/U{row},\"–\")"
                cell.alignment = center; cell.number_format = '0.00%'
            elif col_idx in [3,4,5,7,8,10,11,12,14,15,17,18,19,21,22]:
                cell.alignment = center; cell.number_format = '#,##0'
            else:
                cell.alignment = center

    ws2.freeze_panes = "C3"

    output = "MKTBrs_競品分析工作包.xlsx"
    wb.save(output)
    print(f"\n✅ Excel 已儲存：{output}")


# ── 主程式 ────────────────────────────────────────────────
if __name__ == "__main__":
    print("=== MKTBrs 競品分析工作包產生器 ===\n")

    print("【步驟 1】抓取品牌圖片...")
    logo_paths = {}
    for brand in BRANDS:
        path = fetch_og_image(brand)
        if path:
            logo_paths[brand["name"]] = path

    print("\n【步驟 2】產生 Excel 工作包...")
    build_excel(BRANDS, logo_paths)

    print("\n=== 完成 ===")
    print("產出物：")
    print("  - MKTBrs_競品分析工作包.xlsx")
    print(f"  - logos/ 資料夾（{len(logo_paths)} 張品牌圖片）")
