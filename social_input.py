import openpyxl
from datetime import datetime

EXCEL_PATH = "MKTBrs_競品分析工作包.xlsx"


def ask_int(prompt):
    while True:
        val = input(prompt).strip()
        if val == "":
            return None
        try:
            return int(val.replace(",", ""))
        except ValueError:
            print("   ⚠️  請輸入純數字（例：12500），或直接 Enter 跳過")


def ask_date(prompt):
    while True:
        val = input(prompt).strip()
        if val == "":
            return None
        # 移除所有分隔符號，統一處理
        clean = val.replace("/", "").replace("-", "").replace(".", "")
        try:
            if len(clean) == 8:
                # YYYYMMDD
                return datetime.strptime(clean, "%Y%m%d")
            elif len(clean) == 4:
                # MMDD，補上今年
                year = datetime.now().year
                return datetime.strptime(f"{year}{clean}", "%Y%m%d")
            else:
                print("   ⚠️  格式請輸入 YYYYMMDD 或 MMDD")
        except ValueError:
            print("   ⚠️  日期無效，請重新輸入")


def input_10_posts(platform):
    """輸入近期 10 則貼文，回傳統計結果"""
    print(f"\n  輸入近期 10 則貼文（從最新到最舊）")
    print(f"  日期格式：MM/DD 或 YYYY/MM/DD，不知道可按 Enter 跳過\n")

    dates, likes, comments = [], [], []

    for i in range(1, 11):
        print(f"  第 {i} 則")
        d = ask_date(f"    日期：")
        l = ask_int(f"    讚數：")
        c = ask_int(f"    留言數：")

        if d: dates.append(d)
        if l is not None: likes.append(l)
        if c is not None: comments.append(c)

    # 計算平均
    avg_likes    = round(sum(likes) / len(likes)) if likes else None
    avg_comments = round(sum(comments) / len(comments)) if comments else None

    # 計算發文頻率
    post_frequency = None
    posts_per_month = None
    if len(dates) >= 2:
        dates_sorted = sorted(dates)
        span_days = (dates_sorted[-1] - dates_sorted[0]).days
        if span_days > 0:
            days_per_post = round(span_days / (len(dates) - 1), 1)
            posts_per_month = round(30 / days_per_post, 1)
            post_frequency = f"每 {days_per_post} 天一篇（約每月 {posts_per_month} 篇）"
            print(f"\n  📊 {platform} 發文頻率：{post_frequency}")

    print(f"  📊 {platform} 近期均讚：{avg_likes}｜近期均留言：{avg_comments}")

    return {
        "avg_likes": avg_likes,
        "avg_comments": avg_comments,
        "post_frequency": post_frequency,
        "posts_per_month": posts_per_month,
    }


def input_top_posts(platform):
    """輸入熱門 10 則的平均值"""
    print(f"\n  輸入熱門 10 則平均值（按 Enter 跳過）")
    return {
        "avg_likes":    ask_int(f"  {platform} 熱門均讚："),
        "avg_comments": ask_int(f"  {platform} 熱門均留言："),
    }


def input_brand_data(brand_name):
    print(f"\n{'='*50}")
    print(f"  📋 {brand_name}")
    print(f"{'='*50}")

    # ── Facebook ──
    print("\n【Facebook】")
    fb_fans = ask_int("FB 粉絲數（按 Enter 跳過）：")
    fb_recent = input_10_posts("FB")
    fb_top    = input_top_posts("FB")

    # ── Instagram ──
    print("\n【Instagram】")
    ig_fans = ask_int("IG 粉絲數（按 Enter 跳過）：")
    ig_recent = input_10_posts("IG")
    ig_top    = input_top_posts("IG")

    # ── 商業規模 ──
    print("\n【商業規模】（按 Enter 跳過）")
    biz = {
        "has_ads":   input("有無投放廣告（是／否）：").strip() or None,
        "ad_budget": input("廣告預算規模：").strip() or None,
        "revenue":   input("財報／銷售額：").strip() or None,
        "notes":     input("備註：").strip() or None,
    }

    return {
        "fb_fans":    fb_fans,
        "fb_recent":  fb_recent,
        "fb_top":     fb_top,
        "ig_fans":    ig_fans,
        "ig_recent":  ig_recent,
        "ig_top":     ig_top,
        "biz":        biz,
    }


def write_to_excel(all_data):
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["競品數據追蹤"]

    # 找品牌列
    brand_row_map = {}
    for r in range(3, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val:
            brand_row_map[val] = r

    # 找或建立「發文頻率」欄位
    # 在備註欄（col 27）後面加兩欄：FB發文頻率、IG發文頻率
    # 先檢查 header row 有沒有這兩欄
    last_col = ws.max_column
    freq_col_map = {}
    for c in range(1, last_col + 1):
        val = ws.cell(row=2, column=c).value
        if val in ["FB 發文頻率", "IG 發文頻率"]:
            freq_col_map[val] = c

    if "FB 發文頻率" not in freq_col_map:
        col = last_col + 1
        ws.cell(row=1, column=col).value = "發文頻率"
        ws.cell(row=2, column=col).value = "FB 發文頻率"
        freq_col_map["FB 發文頻率"] = col
        last_col = col

    if "IG 發文頻率" not in freq_col_map:
        col = last_col + 1
        ws.cell(row=2, column=col).value = "IG 發文頻率"
        freq_col_map["IG 發文頻率"] = col

    for brand_name, data in all_data.items():
        r = brand_row_map.get(brand_name)
        if not r:
            print(f"⚠️  找不到品牌列：{brand_name}")
            continue

        fb  = data["fb_recent"]
        fbt = data["fb_top"]
        ig  = data["ig_recent"]
        igt = data["ig_top"]
        biz = data["biz"]

        # FB
        if data["fb_fans"] is not None:
            ws.cell(row=r, column=3).value = data["fb_fans"]
        if fb["avg_likes"] is not None:
            ws.cell(row=r, column=4).value = fb["avg_likes"]
        if fb["avg_comments"] is not None:
            ws.cell(row=r, column=5).value = fb["avg_comments"]
        if fbt["avg_likes"] is not None:
            ws.cell(row=r, column=7).value = fbt["avg_likes"]
        if fbt["avg_comments"] is not None:
            ws.cell(row=r, column=8).value = fbt["avg_comments"]
        if fb["post_frequency"]:
            ws.cell(row=r, column=freq_col_map["FB 發文頻率"]).value = fb["post_frequency"]

        # IG
        if data["ig_fans"] is not None:
            ws.cell(row=r, column=10).value = data["ig_fans"]
        if ig["avg_likes"] is not None:
            ws.cell(row=r, column=11).value = ig["avg_likes"]
        if ig["avg_comments"] is not None:
            ws.cell(row=r, column=12).value = ig["avg_comments"]
        if igt["avg_likes"] is not None:
            ws.cell(row=r, column=14).value = igt["avg_likes"]
        if igt["avg_comments"] is not None:
            ws.cell(row=r, column=15).value = igt["avg_comments"]
        if ig["post_frequency"]:
            ws.cell(row=r, column=freq_col_map["IG 發文頻率"]).value = ig["post_frequency"]

        # 商業規模
        if biz["has_ads"] is not None:
            ws.cell(row=r, column=24).value = biz["has_ads"]
        if biz["ad_budget"] is not None:
            ws.cell(row=r, column=25).value = biz["ad_budget"]
        if biz["revenue"] is not None:
            ws.cell(row=r, column=26).value = biz["revenue"]
        if biz["notes"] is not None:
            ws.cell(row=r, column=27).value = biz["notes"]

        print(f"✅ {brand_name}：數據已寫入")

    wb.save(EXCEL_PATH)
    print(f"\n💾 已儲存：{EXCEL_PATH}")


def main():
    print("\n=== MKTBrs 社群數據輸入介面 ===")
    print("輸入每則貼文數字，程式自動計算平均與發文頻率\n")

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["競品數據追蹤"]

    brands = []
    for r in range(3, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val:
            brands.append(val)

    if not brands:
        print("❌ 找不到品牌資料，請先跑 run.py")
        return

    print(f"找到 {len(brands)} 個品牌：{', '.join(brands)}\n")
    print("要輸入哪個品牌的數據？")
    for i, name in enumerate(brands, start=1):
        print(f"  {i}. {name}")
    print(f"  0. 全部依序輸入")

    choice = input("\n請輸入數字：").strip()

    if choice == "0":
        selected = brands
    elif choice.isdigit() and 1 <= int(choice) <= len(brands):
        selected = [brands[int(choice) - 1]]
    else:
        print("❌ 無效選擇")
        return

    all_data = {}
    for brand_name in selected:
        all_data[brand_name] = input_brand_data(brand_name)

    print("\n── 寫入 Excel ────────────────")
    write_to_excel(all_data)
    print("\n=== 完成 ===")


if __name__ == "__main__":
    main()
